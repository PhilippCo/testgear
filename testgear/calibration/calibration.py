import openpyxl as xls
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from IPython.display import display, HTML

import time
import numpy as np
import testgear


def calibrate(calibrator, dut): 
    filename = "reports/" + dut.calstr.split(",")[0] + time.strftime("_%Y-%m-%d-%H%M%S")+"_v01_cal.xlsx"

    wb = xls.Workbook()
    ws = wb.active

    ws.append(["Calibration Report"])
    ws['A1'].font = Font(size=26, bold=True)

    ws.append([""])
    ws.append(["Cal. Date", time.strftime("%Y-%m-%d")])
    ws.append([""])
    ws.append(["Object", "Digital Multimeter"])
    ws.append(["Manufacturer", dut.idstr.split(",")[0]])
    ws.append(["Type", dut.idstr.split(",")[1]])
    ws.append(["Ser. No.", dut.idstr.split(",")[2]])
    ws.append(["CalString", dut.calstr])
    ws.append(["Inventory", dut.calstr.split(",")[0]])

    ws.append([""])

    ws.append(["Calibration Equipment used:"])
    ws['A12'].font = Font(bold=True)
    ws.append(["Fluke 5730A"])
    ws.append(["Tested against 1-year specicifactions of "+dut.get_class()+" class"])

    ws.append([""])

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 9

    ws.append(["Mode", "True Value", "Frequency", "Range", "Measured", "% of Spec", "Result"])
    ws['A16'].font = Font(bold=True)
    ws['B16'].font = Font(bold=True)
    ws['C16'].font = Font(bold=True)
    ws['D16'].font = Font(bold=True)
    ws['E16'].font = Font(bold=True)
    ws['F16'].font = Font(bold=True)
    ws['G16'].font = Font(bold=True)

    ws['A16'].border = Border(bottom=Side(style='thick'))
    ws['B16'].border = Border(bottom=Side(style='thick'))
    ws['C16'].border = Border(bottom=Side(style='thick'))
    ws['D16'].border = Border(bottom=Side(style='thick'))
    ws['E16'].border = Border(bottom=Side(style='thick'))
    ws['F16'].border = Border(bottom=Side(style='thick'))
    ws['G16'].border = Border(bottom=Side(style='thick'))

    wb.save(filename)
    
    
    
    overall = "PASS"

    for cal in dut.callist:
        print(cal["instruction"])
        input("ENTER when ready")

        if 'calibrator in use' not in cal:
            cal['calibrator in use'] = True
            print("Please update calibration descriton in DMM class")
        

        actual_calpoint = 1
        for calpoint in cal['calpoints']:
            
            print("Step {0}/{1} ".format(actual_calpoint, len(cal['calpoints'])), end="")
            print(calpoint, end="")

            dut.set_mode(calpoint['mode'], calpoint['mrange'])
            limit = dut.get_error(value=calpoint['value'], mode=calpoint['mode'], mrange=calpoint['mrange'], frequency=calpoint['frequency'], calperiod="1 year")

            if cal['calibrator in use'] == True:
                truth = calibrator.set_cal_point(calpoint)
            else:
                truth = {'output': calpoint['value']}
                
            time.sleep(1)

            ###############################
            #10 AVG aktuell
            reading = dut.read_avg(20) 
            ###############################

            dev = np.abs(reading['mean'] - truth['output']) 

            if dev < limit:
                result = "PASS"
            else:
                result  = "FAIL"
                overall = "FAIL"

            perc = round(dev / limit * 100, 0)

            data = [ calpoint['mode'], truth['output'], calpoint['frequency'], calpoint['mrange'], reading['mean'], perc, result]
            
            if result == "PASS":
                display(HTML('<span style="font-weight: bold; color: green;">PASS</span>'))
            else:
                display(HTML('<span style="font-weight: bold; color: red;">FAIL</span>'))
                print("Truth: ", truth['output'])
                print("Reading: ", reading['mean'])
                print("Deviation: ", np.abs(reading['mean'] - truth['output']) )
                print("Limit: ", limit )
                
                
            #print(" ", result)
            
            ws.append(data)
            wb.save(filename)
            actual_calpoint += 1
            
        calibrator.set_output(voltage=0, enabled=False)

    if overall == "PASS":
        dut.write("CALibration:STRing '" + dut.calstr.split(",")[0] + "," +time.strftime("%Y-%m-%d")+ "'" )
        ws['E5'] = "IN SPECIFICATION"
    else:
        ws['E5'] = "OUT OF SPECIFICATION"
    ws['E5'].font = Font(bold=True)
    wb.save(filename)

